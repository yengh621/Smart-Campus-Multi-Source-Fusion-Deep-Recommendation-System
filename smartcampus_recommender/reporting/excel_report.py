"""生成论文可直接复制的实验结果Excel。用户明确要求Python，故使用openpyxl。"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


TASKS = ("knowledge", "course", "consume")
METRICS = ("auc", "ndcg@5", "ndcg@10", "recall@5", "recall@10")


def style_sheet(ws):
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.fill = header_fill; cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    for col in range(1, ws.max_column + 1):
        width = min(max(len(str(ws.cell(r, col).value or "")) for r in range(1, ws.max_row + 1)) + 3, 36)
        ws.column_dimensions[get_column_letter(col)].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, float): cell.number_format = "0.0000"


def write_metric_workbook(path: Path, test_metrics, ablation, config, history,
                          knowledge_mapping_stats=None, formula_parameters=None):
    wb = Workbook(); ws = wb.active; ws.title = "Test Metrics"
    retrieval_key = f"retrieval_recall@{config.retrieval_topk}"
    ws.append(["Task", "Pre-MMR AUC", "Final NDCG@5", "Final NDCG@10", "Final Recall@5", "Final Recall@10",
               f"Retrieval Recall@{config.retrieval_topk}"])
    for task in TASKS:
        ws.append([task, test_metrics[f"{task}_auc"],
                   *[test_metrics[f"{task}_final_{m}"] for m in ("ndcg@5", "ndcg@10", "recall@5", "recall@10")],
                   test_metrics[f"{task}_{retrieval_key}"]])
    ws.conditional_formatting.add(f"B2:G{ws.max_row}", ColorScaleRule(
        start_type="min", start_color="FEE2E2", mid_type="percentile", mid_value=50,
        mid_color="FEF3C7", end_type="max", end_color="DCFCE7"))
    style_sheet(ws)

    wa = wb.create_sheet("Ablation")
    wa.append(["Model", *[f"{task} NDCG@10" for task in TASKS], "Mean NDCG@10"])
    for model, metrics in ablation.items():
        values = [metrics[f"{task}_ndcg@10"] for task in TASKS]
        wa.append([model, *values, sum(values)/len(values)])
    style_sheet(wa)

    wh = wb.create_sheet("Hyperparameters"); wh.append(["Parameter", "Value"])
    for key, value in config.to_dict().items(): wh.append([key, str(value)])
    style_sheet(wh)

    wt = wb.create_sheet("Training History")
    columns = list(history[0]) if history else ["epoch"]
    wt.append(columns)
    for row in history: wt.append([row.get(c) for c in columns])
    style_sheet(wt)

    wn = wb.create_sheet("Protocol")
    wn.append(["Item", "Description"])
    wn.append(["Split", "User-level stratified 70%/15%/15%; no user appears in multiple subsets."])
    wn.append(["Selection", "Best checkpoint selected only by validation mean NDCG@10."])
    wn.append(["Test", "Test set evaluated once after model selection."])
    wn.append(["Identity assumption", "The model treats records sharing global_user_id as one student; the key itself was synthetically constructed."])
    style_sheet(wn)

    wk = wb.create_sheet("Knowledge Mapping")
    wk.append(["Mapping statistic", "Count"])
    for key, value in (knowledge_mapping_stats or {}).items():
        wk.append([key, value])
    style_sheet(wk)

    wd = wb.create_sheet("Diversity & Debias")
    wd.append(["Task", "Coverage", "Gini", "Novelty", "Average Popularity",
               "Intra-list Diversity", "Inter-list Diversity"])
    for task in TASKS:
        wd.append([task, *[test_metrics.get(f"{task}_{name}", "") for name in (
            "coverage", "gini", "novelty", "average_popularity",
            "intra_list_diversity", "inter_list_diversity")]])
    style_sheet(wd)

    wf = wb.create_sheet("Learned Formula Params")
    wf.append(["Parameter group", "Task/Layer", "Head", "Kernel", "Value"])
    formula_parameters = formula_parameters or {}
    for task, value in formula_parameters.get("task_log_variance", {}).items():
        wf.append(["Task log variance", task, "", "", value])
    for task, value in formula_parameters.get("task_normalized_weight", {}).items():
        wf.append(["Normalized task weight", task, "", "", value])
    for layer in formula_parameters.get("time_decay_layers", []):
        for head, (rates, mixtures) in enumerate(zip(layer["rates"], layer["mixtures"]), 1):
            for kernel, (rate, mixture) in enumerate(zip(rates, mixtures), 1):
                wf.append(["Time decay rate", layer["layer"], head, kernel, rate])
                wf.append(["Time mixture weight", layer["layer"], head, kernel, mixture])
    style_sheet(wf)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
